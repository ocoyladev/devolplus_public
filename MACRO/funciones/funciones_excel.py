import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
import math
from pathlib import Path
from decimal import Decimal, InvalidOperation
import re

from ..config import NBSP, CURRENCY_COLS, CURRENCY_EXCEL_FORMAT, SKIP_PERSONALIZADO_PREFIXES
from .funciones_generales import (
    _coerce_to_number, _valor_a_texto, _preparar_registros_para_comparacion,
    _formatear_cambio, _normalizar_texto_para_comparacion, _normalize_numeric_value,
    _is_empty_row, _sum_numeric_values
)

def crear_hoja_cambios(df_base, df_actualizado, ruta_excel_destino):
    """Crea o reemplaza la hoja 'Cambios' en el archivo destino."""
    if df_base is None or df_actualizado is None:
        return

    columna_clave = "num_doc"
    if columna_clave not in df_base.columns or columna_clave not in df_actualizado.columns:
        print("⚠️ No se pudo generar la hoja de cambios porque falta la columna 'num_doc'.")
        return

    columnas_base = list(df_base.columns)
    columnas_actual = list(df_actualizado.columns)
    columnas_ordenadas = list(columnas_base)
    for col in columnas_actual:
        if col not in columnas_ordenadas:
            columnas_ordenadas.append(col)

    df_base_alineado = df_base.copy()
    df_actual_alineado = df_actualizado.copy()
    for col in columnas_ordenadas:
        if col not in df_base_alineado.columns:
            df_base_alineado[col] = pd.NA
        if col not in df_actual_alineado.columns:
            df_actual_alineado[col] = pd.NA

    df_base_alineado = df_base_alineado[columnas_ordenadas]
    df_actual_alineado = df_actual_alineado[columnas_ordenadas]

    registros_base_valores, registros_base_cmp, orden_base = _preparar_registros_para_comparacion(
        df_base_alineado, columnas_ordenadas, columna_clave
    )
    registros_actual_valores, registros_actual_cmp, orden_actual = _preparar_registros_para_comparacion(
        df_actual_alineado, columnas_ordenadas, columna_clave
    )

    orden_total = orden_actual + [
        clave for clave in orden_base if clave not in registros_actual_valores
    ]

    cantidad_columnas_fijas = min(5, len(columnas_ordenadas))
    columnas_fijas = columnas_ordenadas[:cantidad_columnas_fijas]
    columnas_ignorar = {"fec_act"}

    filas = []
    banderas_cambio = []
    columnas_con_cambios = set()

    for clave in orden_total:
        base_row_valores = registros_base_valores.get(
            clave, {col: "" for col in columnas_ordenadas}
        )
        actual_row_valores = registros_actual_valores.get(
            clave, {col: "" for col in columnas_ordenadas}
        )
        base_row_cmp = registros_base_cmp.get(
            clave, {col: "" for col in columnas_ordenadas}
        )
        actual_row_cmp = registros_actual_cmp.get(
            clave, {col: "" for col in columnas_ordenadas}
        )

        columnas_diferentes = []
        for col in columnas_ordenadas:
            if col in columnas_ignorar:
                continue
            if base_row_cmp.get(col, "") != actual_row_cmp.get(col, ""):
                columnas_diferentes.append(col)

        if not columnas_diferentes:
            continue

        fila_valores = {
            col: actual_row_valores.get(col, "") for col in columnas_ordenadas
        }
        fila_banderas = {col: False for col in columnas_ordenadas}

        for col in columnas_diferentes:
            fila_valores[col] = _formatear_cambio(
                base_row_valores.get(col, ""), actual_row_valores.get(col, "")
            )
            fila_banderas[col] = True
            if col not in columnas_fijas:
                columnas_con_cambios.add(col)

        filas.append(fila_valores)
        banderas_cambio.append(fila_banderas)

    columnas_cambios_ordenadas = [
        col for col in columnas_ordenadas if col in columnas_con_cambios
    ]
    columnas_salida = columnas_fijas + columnas_cambios_ordenadas
    if not columnas_salida:
        columnas_salida = [columna_clave]

    try:
        wb = load_workbook(ruta_excel_destino)
    except Exception as e:
        print(f"⚠️ No se pudo abrir '{ruta_excel_destino}' para agregar la hoja de cambios: {e}")
        return

    if "Cambios" in wb.sheetnames:
        del wb["Cambios"]

    ws = wb.create_sheet(title="Cambios", index=1 if wb.sheetnames else 0)
    ws.freeze_panes = "A2"

    header_font = Font(bold=True)
    cambio_fill = PatternFill(fill_type="solid", fgColor="FFF4B084")

    for idx, col in enumerate(columnas_salida, start=1):
        ws.cell(row=1, column=idx, value=col).font = header_font

    if not filas:
        mensaje = "No se detectaron cambios entre el archivo base y los datos actualizados."
        ws.cell(row=2, column=1, value=mensaje)
        if len(columnas_salida) > 1:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columnas_salida))
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
        autosize_columns(ws)
        wb.save(ruta_excel_destino)
        wb.close()
        return

    for row_idx, fila in enumerate(filas, start=2):
        banderas = banderas_cambio[row_idx - 2]
        for col_idx, col in enumerate(columnas_salida, start=1):
            valor = fila.get(col, "")
            celda = ws.cell(row=row_idx, column=col_idx, value=valor)
            if banderas.get(col):
                celda.fill = cambio_fill

    autosize_columns(ws)
    wb.save(ruta_excel_destino)
    wb.close()

def preparar_hoja_con_titulo(ws, titulo, num_cols):
    """
    Reserva fila 1 para título y fila 2 en blanco.
    No inserta filas una vez creada la tabla (evita dañar tablas).
    Devuelve la fila donde debe empezar el header de datos (3).
    """
    num_cols = max(1, int(num_cols or 1))
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = titulo
    c.font = Font(bold=True, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")
    # asegura dos filas ocupadas arriba
    while ws.max_row < 2:
        ws.append([""])  # mejor que [], para que cuente la fila

    return 3

def _first_nonempty_row_as_header(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Toma un DataFrame leído sin encabezados (header=None) y usa la primera
    fila no vacía como encabezado. Devuelve el DF limpio, sin filas/columnas vacías.
    """
    if df_raw.empty:
        return df_raw

    # índice de la fila con al menos 3 celdas no vacías (heurística):
    header_idx = None
    for i, row in df_raw.iterrows():
        if row.notna().sum() >= 3:
            header_idx = i
            break
    if header_idx is None:
        header_idx = 0

    headers = [
        (str(v).strip() if pd.notna(v) and str(v).strip() not in ("", "nan", "None") else f"Columna{i+1}")
        for i, v in enumerate(df_raw.iloc[header_idx].tolist())
    ]
    df = df_raw.iloc[header_idx + 1:].copy()
    df.columns = headers
    # limpiar filas/columnas totalmente vacías
    df = df.dropna(how="all")
    df = df.loc[:, df.notna().any(axis=0)]
    # normalizar espacios
    df.columns = [str(c).strip() for c in df.columns]
    return df

def append_dataframe_to_sheet(ws, df):
    for row in dataframe_to_rows(df, index=False, header=True):
        if _is_empty_row(row):
            continue
        ws.append(list(row))

def autosize_columns(ws):
    for col_idx, column_cells in enumerate(ws.columns, 1):
        length = max(len(str(c.value)) if c.value else 0 for c in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, length+2), 50)

def autosize_specific_column(ws, column_letter, min_row=1, max_row=None, min_width=10, max_width=50):
    try:
        col_idx = column_index_from_string(column_letter)
    except ValueError:
        return

    if col_idx > ws.max_column:
        return

    if max_row is None or max_row > ws.max_row:
        max_row = ws.max_row

    max_length = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=col_idx, max_col=col_idx):
        cell = row[0]
        if cell.value is None:
            continue
        length = len(str(cell.value))
        if length > max_length:
            max_length = length

    width = min(max(min_width, max_length + 2), max_width) if max_length else min_width
    ws.column_dimensions[column_letter].width = width

def _unique_table_name(wb, base_name):
    existing = set()
    for sheet in wb.worksheets:
        existing.update(sheet.tables.keys())

    candidate = base_name
    counter = 1
    while candidate in existing:
        candidate = f"{base_name}_{counter}"
        counter += 1
    return candidate

def format_sheet_generic(
    wb,
    ws,
    df,
    table_base_name,
    *,
    currency_letters=(),
    sum_letters=(),
    auto_columns=(),
    fixed_widths=None,
    auto_width_all=False,
    total_label="Total",
    header_row=1 
):
    if df is None or df.empty:
        return

    nrows, ncols = df.shape
    # header_row = 1
    data_start_row = header_row + 1
    last_data_row = nrows + header_row
    total_row_idx = last_data_row + 1

    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=header_row, max_row=last_data_row, min_col=1, max_col=ncols):
        for cell in row:
            cell.alignment = center_alignment
            if cell.row == header_row:
                cell.font = header_font

    for row_idx in range(header_row, total_row_idx + 1):
        ws.row_dimensions[row_idx].height = 30

    currency_letters = [
        letter for letter in {letter.upper() for letter in currency_letters}
        if column_index_from_string(letter) <= ncols
    ]

    sum_letters = {
        letter.upper() for letter in sum_letters
        if column_index_from_string(letter) <= ncols
    }

    for letter in currency_letters:
        col_idx = column_index_from_string(letter)
        for row_idx in range(data_start_row, last_data_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            numeric_value = _coerce_to_number(cell.value)
            if numeric_value is not None:
                cell.value = numeric_value
            cell.number_format = CURRENCY_EXCEL_FORMAT
            cell.alignment = center_alignment
        ws.cell(row=total_row_idx, column=col_idx).number_format = CURRENCY_EXCEL_FORMAT
        sum_letters.add(letter)

    table_name = _unique_table_name(wb, table_base_name)
    ref = f"A{header_row}:{get_column_letter(ncols)}{total_row_idx}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight8", showRowStripes=True)
    table.showTotalsRow = True
    table.tableColumns = []

    for idx, column_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(idx)
        total_cell = ws.cell(row=total_row_idx, column=idx)
        total_cell.alignment = center_alignment

        tc = TableColumn(id=idx, name=str(column_name) if column_name not in (None, "") else f"Col_{idx}")

        if idx == 1:
            tc.totalsRowLabel = total_label
            total_cell.value = total_label
            total_cell.font = Font(bold=True)
        elif col_letter in sum_letters or (not sum_letters and idx == ncols):
            tc.totalsRowFunction = "sum"
            total_cell.value = (
                f"=SUBTOTAL(109,{col_letter}{data_start_row}:{col_letter}{last_data_row})"
            )
            if col_letter in currency_letters:
                total_cell.number_format = CURRENCY_EXCEL_FORMAT

        table.tableColumns.append(tc)

    ws.add_table(table)

    if auto_columns:
        for letter in {letter.upper() for letter in auto_columns}:
            if column_index_from_string(letter) <= ncols:
                autosize_specific_column(ws, letter, max_row=total_row_idx)

    if fixed_widths:
        for letter, width in fixed_widths.items():
            try:
                if column_index_from_string(letter) <= ncols:
                    ws.column_dimensions[letter.upper()].width = width
            except ValueError:
                continue

    if auto_width_all:
        for idx in range(1, ncols + 1):
            autosize_specific_column(ws, get_column_letter(idx), max_row=total_row_idx)

def write_df_as_table(ws, df, start_row, table_name):
    if df.empty or df.shape[1] == 0:
        print(f"[WARN] DataFrame vacío en {table_name}, se omite.")
        return start_row

    # Asegurar nombre de tabla único en todo el libro
    table_name = _unique_table_name(ws.parent, table_name)
    # Normalizar encabezados
    df.columns = [
        str(c) if c not in ("", None, "nan") else f"Col_{i+1}"
        for i, c in enumerate(df.columns)
    ]

    # Encabezados en negrita y color blanco
    header_font = Font(bold=True, color="FFFFFF")
    for j, col in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=j, value=col).font = header_font

    # Datos
    for i in range(df.shape[0]):
        for j in range(df.shape[1]):
            val = df.iat[i, j]
            cell = ws.cell(row=start_row + 1 + i, column=j + 1, value=val)
            if df.columns[j] in CURRENCY_COLS:
                cell.number_format = CURRENCY_EXCEL_FORMAT

    # Fila de totales para la tabla
    total_row = start_row + df.shape[0] + 1

    # Crear tabla con estilo y fila de total
    ref = f"A{start_row}:{get_column_letter(df.shape[1])}{total_row}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleLight8", showRowStripes=True)
    table.tableStyleInfo = style
    table.showTotalsRow = True

    # Definir columnas y totales
    table.tableColumns = []
    for idx, colname in enumerate(df.columns, 1):
        tc = TableColumn(id=idx, name=colname)

        cell = ws.cell(row=total_row, column=idx)

        if idx == 1:
            tc.totalsRowLabel = "Total"
            cell.value = "Total"
            cell.font = Font(bold=True)

        if colname in CURRENCY_COLS:
            tc.totalsRowFunction = "sum"
            col_letter = get_column_letter(idx)
            cell.value = f"=SUBTOTAL(109,{col_letter}{start_row+1}:{col_letter}{total_row-1})"
            cell.number_format = CURRENCY_EXCEL_FORMAT

        table.tableColumns.append(tc)

    ws.add_table(table)

    return total_row + 3

def _sanitize_personalizado_dataframe(df):
    """Devuelve una copia del DataFrame sin filas de totales preexistentes."""
    if df.empty:
        return df.copy()

    df_clean = df.copy()
    primera_columna = df_clean.columns[0] if df_clean.columns.size else None
    if primera_columna:
        valores = (
            df_clean[primera_columna]
            .astype(str)
            .str.strip()
            .str.upper()
        )
        df_clean = df_clean.loc[~valores.isin({"TOTAL", "TOTALES"})]
    return df_clean

def _sumar_columnas(df, candidatos):
    """Suma la primera columna encontrada dentro de la lista de candidatos."""
    if df.empty:
        return 0.0

    columnas_normalizadas = {
        str(col).strip().upper(): col for col in df.columns if col is not None
    }

    for candidato in candidatos:
        nombre_normalizado = candidato.strip().upper()
        if nombre_normalizado in columnas_normalizadas:
            col_real = columnas_normalizadas[nombre_normalizado]
            serie = pd.to_numeric(df[col_real], errors="coerce").fillna(0)
            return float(serie.sum())

    return 0.0

def actualizar_resumen_personalizado(resumen, nombre_archivo, df):
    """Actualiza el diccionario de resumen con los totales del DataFrame."""
    if df.empty:
        return

    df_clean = _sanitize_personalizado_dataframe(df)
    if df_clean.empty:
        return

    nombre = Path(nombre_archivo).stem.lower()

    if nombre.startswith("pag_4ta_"):
        resumen["Pag_4ta"] += _sumar_columnas(df_clean, ["PAGO TOTAL", "PAGOS SIN INTERESES"])
    elif nombre.startswith("rbr_4ta_"):
        resumen["Rta_4ta_107"] += _sumar_columnas(df_clean, ["INGRESO ATRIBUIDO"])
    elif nombre.startswith("otr_4ta_"):
        # Sumar la última columna para Rentas de Cuarta Categoría (Otras Rentas)
        if not df_clean.empty:
            last_col = df_clean.columns[-1]
            resumen["Rta_4ta_108"] += _sumar_columnas(df_clean, [last_col])
    elif nombre.startswith("rbr_5ta_"):
        resumen["Rta_5ta"] += _sumar_columnas(df_clean, ["INGRESO ATRIBUIDO", "REMUNERACION MENSUAL"])
    elif nombre.startswith("ret_4ta_"):
        resumen["Ret_4ta"] += _sumar_columnas(
            df_clean,
            [
                "RETENCION",
                "RETENCION POR RECIBOS, DIETAS, OTROS COMPROBANTES Y NOTAS DE CREDITO",
            ],
        )
    elif nombre.startswith("ret_5ta_"):
        retenciones = _sumar_columnas(df_clean, ["RETENCION"])
        devoluciones = _sumar_columnas(df_clean, ["DEVOLUCION POR EXCESO DE RETENCION"])
        resumen["Ret_5ta"] += retenciones - devoluciones
    elif nombre.startswith("ret_itf_"):
        resumen["ITF"] += _sumar_columnas(df_clean, ["MONTO DE LA RETENCION", "RETENCION"])

def colocar_resumen_en_plantilla(wb, resumen, sheet_name="RELIQUIDADO", col_idx=7):
    """Ubica los totales obtenidos en la hoja principal de la plantilla."""
    if sheet_name not in wb.sheetnames:
        print(f"[WARN] El libro no contiene la hoja '{sheet_name}'.")
        return

    ws_destino = wb[sheet_name]

    # Mapeo de filas (la columna es parametrizable, G es 7)
    mapeo_filas = {
        "Rta_4ta_107": 15,
        "Rta_4ta_108": 17,
        "Rta_5ta": 19,
        "ITF": 29,
        "Ret_4ta": 39,
        "Ret_5ta": 40,
        "Pag_4ta": 43,
    }

    for clave, fila in mapeo_filas.items():
        valor = resumen.get(clave, 0.0)
        if isinstance(valor, (int, float)) and not math.isfinite(valor):
            valor = 0.0
        if isinstance(valor, float):
            # Si no es RELIQUIDADO (es decir, es RELIQ-ALL u otro), redondear sin decimales
            # if sheet_name != "RELIQUIDADO":
            valor = round(valor)
            # else:
            #     valor = round(valor, 2)
        
        ws_destino.cell(row=fila, column=col_idx, value=valor)

def limpiar_texto(s):
    """Normaliza valores de celda preservando espacios internos."""
    return str(s).replace(NBSP, " ").replace("'", "").strip() if s is not None else None

def normalizar_header(h):
    return str(h).replace(NBSP," ").strip() if h else ""

def read_excel_safely(path):
    """Lee un archivo Excel en un DataFrame."""
    ext = Path(path).suffix.lower()
    engines = ["openpyxl"]
    if ext == ".xls":
        engines = ["xlrd", "openpyxl"]  # intentar con xlrd y como respaldo openpyxl

    last_exc = None
    for eng in engines:
        try:
            return pd.read_excel(path, header=0, dtype=str, engine=eng)
        except Exception as e:
            last_exc = e

    print(f"[ERROR] No se pudo leer {path}: {last_exc}")
    return pd.DataFrame()

def extract_year_from_value(val):
    s = limpiar_texto(val)
    if not s: return None
    m = re.search(r"(20\d{2}|19\d{2})", s)
    return m.group(1) if m else None

def titulo_bloque(nombre, anio):
    nombre = nombre.lower()
    if nombre.startswith("otr_4ta_"): return f"RENTAS DE CUARTA CATEGORÍA (OTRAS RENTAS) - {anio}"
    if nombre.startswith("pag_1ra_"): return f"PAGOS DIRECTOS DE PRIMERA CATEGORÍA - {anio}"
    if nombre.startswith("pag_4ta_"): return f"PAGOS DIRECTOS DE CUARTA CATEGORÍA - {anio}"
    if nombre.startswith("pag_5ta_"): return f"PAGOS DIRECTOS DE QUINTA CATEGORÍA - {anio}"
    if nombre.startswith("rbr_2da_"): return f"RENTAS DE SEGUNDA CATEGORÍA - {anio}"
    if nombre.startswith("rbr_4ta_"): return f"RENTAS DE CUARTA CATEGORÍA - {anio}"
    if nombre.startswith("rbr_5ta_"): return f"RENTAS DE QUINTA CATEGORÍA - {anio}"
    if nombre.startswith("ret_2da_"): return f"RETENCIONES DE SEGUNDA CATEGORÍA - {anio}"
    if nombre.startswith("ret_4ta_"): return f"RETENCIONES DE CUARTA CATEGORÍA - {anio}"
    if nombre.startswith("ret_5ta_"): return f"RETENCIONES DE QUINTA CATEGORÍA - {anio}"
    if nombre.startswith("ret_itf_"): return f"ITF - {anio}"
    if nombre.startswith("sal_fav_"): return f"SALDOS A FAVOR - {anio}"
    return None

def procesar_dataframe(df, nombre):
    if df.empty or df.shape[1] == 0:
        return pd.DataFrame()
    df.columns = [normalizar_header(c) for c in df.columns]
    if re.match(r"^(OtR_4ta_|RBr_4ta_|RBr_5ta_|Ret_4ta_|Ret_5ta_|Ret_ITF_)", nombre):
        if df.shape[1] > 0:
            df = df.drop(df.columns[0], axis=1)

    # Normalizar valores; usar DataFrame.map si está disponible (pandas >=2.1)
    func = lambda x: limpiar_texto(x) if pd.notna(x) else x
    try:
        df = df.map(func)  # type: ignore[attr-defined]
    except AttributeError:  # pandas <2.1
        df = df.applymap(func)
    if nombre.startswith(("Pag_4ta_", "Pag_5ta_")):
        for idx in (6, 5, 4):
            if df.shape[1] > idx and df[df.columns[idx]].dropna().shape[0] <= 1:
                df = df.drop(df.columns[idx], axis=1)
    for col in df.columns:
        if col in CURRENCY_COLS:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    if nombre.startswith("Ret_5ta_"):
        numeric = df.apply(pd.to_numeric, errors="coerce")
        zero_cols = numeric.notna().any() & numeric.fillna(0).eq(0).all()
        df = df.loc[:, ~zero_cols]
        df = df.replace("", pd.NA).dropna(axis=1, how="all")
    return df

def guardar_en_excel_con_tabla(df, nombre_archivo='Reporte_nueva_data.xlsx'):
    """Guarda el DataFrame en un archivo Excel con formato de tabla, 
    cabecera coloreada según diseño y columna Auditor responsable como segunda columna."""

    if df is None or df.empty:
        print("No hay datos para guardar en el archivo Excel.")
        return

    try:
        # --- 1. Insertar columna Auditor responsable como segunda columna ---
        cols = list(df.columns)
        # Si la primera columna es "of_devolucion" (OF devolución)
        if cols[0].lower() in ["of_devolucion", "of devolución"]:
            nuevas_cols = [cols[0], "Auditor responsable"] + cols[1:]
        else:
            nuevas_cols = ["Auditor responsable"] + cols
        
        df["Auditor responsable"] = ""  # crear la columna vacía
        df = df.reindex(columns=nuevas_cols)  # reordenar columnas

        # --- 2. Exportar a Excel con formato ---
        writer = pd.ExcelWriter(nombre_archivo, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Datos Filtrados', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Datos Filtrados']

        # --- 3. Colores de cabeceras por rango personalizado ---
        color_ranges = [
            (1, 19, "#2F75B5"),
            (20, 41, "#BF8F00"),
            (42, 58, "#833C0C"),
            (59, 108, "#7B7B7B"),
            (109, 110, "#70AD47"),
            (111, 115, "#FF0000"),
            (116, 151, "#8EA9DB"),
            (152, 165, "#0070C0"),
            (166, 166, "#70AD47"),
            (167, 168, "#FFFF00"),
            (169, 175, "#00B0F0"),
            (176, 181, "#FFD966"),
            (182, 182, "#7030A0"),
        ]

        # Generar lista de colores para cada columna
        num_cols = len(df.columns)
        colores = []
        for i in range(1, num_cols + 1):
            color = "#D9D9D9"  # color por defecto
            for start, end, c in color_ranges:
                if start <= i <= end:
                    color = c
                    break
            colores.append(color)

        formatos = [
            workbook.add_format({
                'bold': True,
                'bg_color': color,
                'font_color': 'white',
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'text_wrap': True
            })
            for color in colores
        ]
        
        # --- 4. Escribir cabeceras con colores ---
        worksheet.set_row(0, 60)  # Alto de fila de encabezado
        for col_num, fmt in enumerate(formatos):
            worksheet.write(0, col_num, df.columns[col_num], fmt)

        # --- 5. Formato de tabla ---
        (max_row, max_col) = df.shape
        column_settings = [{'header': col} for col in df.columns]
        worksheet.add_table(0, 0, max_row, max_col - 1,
                            {'columns': column_settings, 'style': 'Table Style Light 9'})

        # --- 6. Autoajustar ancho ---
        for i in range(len(df.columns)):
            worksheet.set_column(i, i, 12)

        # --- 7. Sobrescribir cabeceras con formato personalizado ---
        for col_num, fmt in enumerate(formatos):
            worksheet.write(0, col_num, df.columns[col_num], fmt)

        writer.close()
        print(f"Proceso finalizado. Archivo '{nombre_archivo}' generado con éxito.")

    except Exception as e:
        print(f"❌ No se pudo guardar el archivo Excel. Error: {e}")
